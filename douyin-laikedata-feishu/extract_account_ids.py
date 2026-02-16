#!/usr/bin/env python3
"""
从 Excel 文件提取账户ID
"""
import sys

# 尝试导入 openpyxl
try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ 缺少 openpyxl 库")
    print("请安装: pip install openpyxl")
    sys.exit(1)

file_path = '/root/单元投放_账户列表_64763_2026_02_13 00_57_23.xlsx'

print("="*60)
print("读取账户列表 Excel")
print("="*60 + "\n")

try:
    # 加载工作簿
    wb = load_workbook(file_path, read_only=True)
    
    # 获取第一个工作表
    ws = wb.active
    sheet_name = wb.sheetnames[0]
    
    print(f"工作表名称: {sheet_name}")
    print(f"总行数: {ws.max_row}")
    print(f"总列数: {ws.max_column}\n")
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    print(f"列名: {headers}\n")
    
    # 显示前10行数据
    print("前10行数据:")
    print("-"*60)
    
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=11, values_only=True), 1):
        print(f"第{i}行: {row}")
    
    print("\n" + "="*60)
    
    # 查找包含 ID 的列
    id_col_index = None
    for i, header in enumerate(headers):
        if header and ('id' in str(header).lower() or 'ID' in str(header) or '账户' in str(header)):
            id_col_index = i
            print(f"找到可能的ID列: 第{i+1}列 ({header})")
    
    if id_col_index is not None:
        print(f"\n提取第{id_col_index+1}列的所有数据...")
        
        account_ids = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
            value = row[id_col_index]
            if value:
                # 尝试转换为整数
                try:
                    account_id = int(value)
                    account_ids.append(account_id)
                except (ValueError, TypeError):
                    # 如果不是数字，尝试提取数字部分
                    import re
                    numbers = re.findall(r'\d+', str(value))
                    if numbers:
                        account_id = int(numbers[0])
                        account_ids.append(account_id)
        
        print(f"\n✅ 提取到 {len(account_ids)} 个账户ID")
        print(f"\n账户ID列表:")
        for acc_id in account_ids[:20]:  # 显示前20个
            print(f"  - {acc_id}")
        
        if len(account_ids) > 20:
            print(f"  ... 还有 {len(account_ids) - 20} 个")
        
        # 保存到文件
        output_file = '/root/.openclaw/workspace/douyin-laikedata-feishu/account_ids.txt'
        with open(output_file, 'w') as f:
            for acc_id in account_ids:
                f.write(f"{acc_id}\n")
        
        print(f"\n💾 已保存到: {output_file}")
    else:
        print("\n⚠️  未找到明确的ID列，请手动指定")
    
    wb.close()
    
except Exception as e:
    print(f"❌ 错误: {e}")
    import traceback
    traceback.print_exc()
